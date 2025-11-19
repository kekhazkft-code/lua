"""
Report Generator for Test Results
Generates Excel/CSV reports with test results
"""

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import List
from dataclasses import asdict

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from test_runner import TestResult


class ReportGenerator:
    """Generates reports from test results"""

    def __init__(self, results: List[TestResult]):
        self.results = results
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    def generate_csv(self, output_path: str = None) -> str:
        """
        Generate CSV report
        Returns path to generated file
        """
        if output_path is None:
            output_path = f"test_results_{self.timestamp}.csv"

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Header
            writer.writerow([
                'Test ID',
                'Category',
                'Description',
                'Initial State',
                'Inputs',
                'Expected Output',
                'Actual Output',
                'Status',
                'Error Message',
                'Execution Time (ms)',
                'Propagation Count',
                'Blocked Count',
                'Relay States',
                'Notes'
            ])

            # Data rows
            for result in self.results:
                writer.writerow([
                    result.test_id,
                    result.category,
                    result.description,
                    result.initial_state,
                    result.inputs,
                    result.expected_output,
                    result.actual_output,
                    'PASS' if result.passed else 'FAIL',
                    result.error_message,
                    f"{result.execution_time_ms:.2f}",
                    result.propagation_count,
                    result.blocked_count,
                    result.relay_states,
                    result.notes
                ])

        print(f"CSV report generated: {output_path}")
        return str(output_path)

    def generate_excel(self, output_path: str = None) -> str:
        """
        Generate Excel report with formatting
        Returns path to generated file
        """
        if not HAS_OPENPYXL:
            print("Warning: openpyxl not available, generating CSV instead")
            return self.generate_csv(output_path)

        if output_path is None:
            output_path = f"test_results_{self.timestamp}.xlsx"

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = openpyxl.Workbook()

        # Summary sheet
        self._create_summary_sheet(wb)

        # Detailed results sheet
        self._create_detailed_sheet(wb)

        # Category breakdown sheets
        self._create_category_sheets(wb)

        # Save workbook
        wb.save(output_path)

        print(f"Excel report generated: {output_path}")
        return str(output_path)

    def _create_summary_sheet(self, wb):
        """Create summary sheet with overall statistics"""
        if 'Sheet' in wb.sheetnames:
            ws = wb['Sheet']
            ws.title = 'Summary'
        else:
            ws = wb.create_sheet('Summary', 0)

        # Title
        ws['A1'] = 'Aging Chamber Test Results - Summary'
        ws['A1'].font = Font(size=16, bold=True)

        # Test run info
        ws['A3'] = 'Test Run Date:'
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        ws['A4'] = 'Total Tests:'
        ws['B4'] = len(self.results)

        passed = sum(1 for r in self.results if r.passed)
        failed = len(self.results) - passed

        ws['A5'] = 'Passed:'
        ws['B5'] = passed
        ws['B5'].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        ws['A6'] = 'Failed:'
        ws['B6'] = failed
        if failed > 0:
            ws['B6'].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

        ws['A7'] = 'Pass Rate:'
        ws['B7'] = f"{passed/len(self.results)*100:.1f}%"

        # Average execution time
        avg_time = sum(r.execution_time_ms for r in self.results) / len(self.results)
        ws['A8'] = 'Avg Execution Time:'
        ws['B8'] = f"{avg_time:.2f} ms"

        # Total propagations
        total_props = sum(r.propagation_count for r in self.results)
        total_blocked = sum(r.blocked_count for r in self.results)

        ws['A10'] = 'Total Propagations:'
        ws['B10'] = total_props

        ws['A11'] = 'Total Blocked:'
        ws['B11'] = total_blocked

        ws['A12'] = 'Block Rate:'
        if total_props + total_blocked > 0:
            ws['B12'] = f"{total_blocked/(total_props + total_blocked)*100:.1f}%"

        # Category breakdown
        ws['A14'] = 'Results by Category'
        ws['A14'].font = Font(size=14, bold=True)

        ws['A15'] = 'Category'
        ws['B15'] = 'Total'
        ws['C15'] = 'Passed'
        ws['D15'] = 'Failed'
        ws['E15'] = 'Pass Rate'

        # Make header bold
        for col in ['A15', 'B15', 'C15', 'D15', 'E15']:
            ws[col].font = Font(bold=True)

        categories = {}
        for result in self.results:
            cat = result.category
            if cat not in categories:
                categories[cat] = {'total': 0, 'passed': 0, 'failed': 0}
            categories[cat]['total'] += 1
            if result.passed:
                categories[cat]['passed'] += 1
            else:
                categories[cat]['failed'] += 1

        row = 16
        for cat, stats in sorted(categories.items()):
            ws[f'A{row}'] = cat
            ws[f'B{row}'] = stats['total']
            ws[f'C{row}'] = stats['passed']
            ws[f'D{row}'] = stats['failed']
            ws[f'E{row}'] = f"{stats['passed']/stats['total']*100:.1f}%"
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

    def _create_detailed_sheet(self, wb):
        """Create detailed results sheet"""
        ws = wb.create_sheet('Detailed Results')

        # Headers
        headers = [
            'Test ID', 'Category', 'Description', 'Status',
            'Execution Time (ms)', 'Propagation Count', 'Blocked Count',
            'Initial State', 'Inputs', 'Expected', 'Actual',
            'Error Message', 'Relay States', 'Notes'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Data rows
        for row, result in enumerate(self.results, 2):
            ws.cell(row=row, column=1).value = result.test_id
            ws.cell(row=row, column=2).value = result.category
            ws.cell(row=row, column=3).value = result.description

            status_cell = ws.cell(row=row, column=4)
            status_cell.value = 'PASS' if result.passed else 'FAIL'
            if result.passed:
                status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

            ws.cell(row=row, column=5).value = f"{result.execution_time_ms:.2f}"
            ws.cell(row=row, column=6).value = result.propagation_count
            ws.cell(row=row, column=7).value = result.blocked_count
            ws.cell(row=row, column=8).value = result.initial_state
            ws.cell(row=row, column=9).value = result.inputs
            ws.cell(row=row, column=10).value = result.expected_output
            ws.cell(row=row, column=11).value = result.actual_output
            ws.cell(row=row, column=12).value = result.error_message
            ws.cell(row=row, column=13).value = result.relay_states
            ws.cell(row=row, column=14).value = result.notes

        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15

        # Freeze header row
        ws.freeze_panes = 'A2'

    def _create_category_sheets(self, wb):
        """Create separate sheet for each test category"""
        categories = {}
        for result in self.results:
            cat = result.category
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(result)

        for category, results in categories.items():
            # Clean sheet name (Excel has restrictions)
            sheet_name = category[:31]  # Max 31 chars
            ws = wb.create_sheet(sheet_name)

            # Headers
            headers = [
                'Test ID', 'Description', 'Status',
                'Execution Time', 'Propagations', 'Blocked',
                'Expected', 'Actual', 'Error'
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)

            # Data
            for row, result in enumerate(results, 2):
                ws.cell(row=row, column=1).value = result.test_id
                ws.cell(row=row, column=2).value = result.description

                status_cell = ws.cell(row=row, column=3)
                status_cell.value = 'PASS' if result.passed else 'FAIL'
                if result.passed:
                    status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

                ws.cell(row=row, column=4).value = f"{result.execution_time_ms:.2f} ms"
                ws.cell(row=row, column=5).value = result.propagation_count
                ws.cell(row=row, column=6).value = result.blocked_count
                ws.cell(row=row, column=7).value = result.expected_output
                ws.cell(row=row, column=8).value = result.actual_output
                ws.cell(row=row, column=9).value = result.error_message

            # Adjust columns
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 10

    def generate_json(self, output_path: str = None) -> str:
        """
        Generate JSON report
        Returns path to generated file
        """
        if output_path is None:
            output_path = f"test_results_{self.timestamp}.json"

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        data = {
            'timestamp': datetime.now().isoformat(),
            'total_tests': len(self.results),
            'passed': sum(1 for r in self.results if r.passed),
            'failed': sum(1 for r in self.results if not r.passed),
            'results': [asdict(r) for r in self.results]
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)

        print(f"JSON report generated: {output_path}")
        return str(output_path)


def main():
    """Test report generation"""
    from test_runner import LuaTestRunner
    from scenario_generator import ScenarioGenerator

    # Generate test data
    generator = ScenarioGenerator()
    scenarios = generator.generate_all_scenarios()[:10]  # Just first 10 for testing

    runner = LuaTestRunner("../aging_chamber_Apar2_0_REFACTORED.lua")
    results = runner.run_all_tests(scenarios)

    # Generate reports
    report_gen = ReportGenerator(results)
    report_gen.generate_csv("../test_reports/results.csv")
    report_gen.generate_excel("../test_reports/results.xlsx")
    report_gen.generate_json("../test_reports/results.json")


if __name__ == '__main__':
    main()
